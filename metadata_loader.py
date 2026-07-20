"""
metadata_loader.py — Load PK/FK metadata repository from Excel or SQL.

Reads Primary Key and Foreign Key constraints from `Primary-Foreign keys.xlsx`
or `Primary-Foreign keys.sql` in `input/ddl/` and exposes a unified lookup model
for DDL conversion and Data Generation.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


@dataclass
class PKConstraint:
    module_name    : str
    table_name     : str
    column_name    : str
    constraint_name: str


@dataclass
class FKConstraint:
    module_name  : str
    child_table  : str
    child_column : str
    parent_table : str
    parent_column: str
    fk_name      : str
    pk_name      : str = ""
    position_no  : int = 1


class MetadataLoader:
    """Loads and caches PK/FK metadata from Excel or SQL sources."""

    def __init__(self, ddl_dir: Path) -> None:
        self.ddl_dir = ddl_dir
        self.pks: dict[str, list[PKConstraint]] = {}  # lowercase table_name -> list of PKConstraint
        self.fks: dict[str, list[FKConstraint]] = {}  # lowercase child_table -> list of FKConstraint
        self._loaded = False

    def load(self) -> None:
        """Load metadata if not already loaded."""
        if self._loaded:
            return

        excel_path = self.ddl_dir / "Primary-Foreign keys.xlsx"
        sql_path   = self.ddl_dir / "Primary-Foreign keys.sql"

        if excel_path.exists():
            self._load_from_excel(excel_path)
        elif sql_path.exists():
            self._load_from_sql(sql_path)
        else:
            logger.info("No PK/FK metadata file found in %s", self.ddl_dir)

        self._loaded = True

    def get_pks_for_table(self, table_name: str) -> list[PKConstraint]:
        self.load()
        return self.pks.get(table_name.lower(), [])

    def get_fks_for_table(self, table_name: str) -> list[FKConstraint]:
        self.load()
        return self.fks.get(table_name.lower(), [])

    def get_all_fks(self) -> list[FKConstraint]:
        self.load()
        all_fks = []
        for fk_list in self.fks.values():
            all_fks.extend(fk_list)
        return all_fks

    def _load_from_excel(self, excel_path: Path) -> None:
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed; skipping Excel metadata load.")
            return

        logger.info("Loading PK/FK metadata from Excel: %s", excel_path.name)
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Load Sheet 1: Foreign Keys
        if "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(c).upper() if c else "" for c in rows[0]]
                idx_mod = self._find_idx(header, ["MODULE", "MODULE_NAME"])
                idx_ct  = self._find_idx(header, ["CHILD_TABLE", "CHILD_TBL"])
                idx_cc  = self._find_idx(header, ["CHILD_COLUMN", "CHILD_COL"])
                idx_pt  = self._find_idx(header, ["PARENT_TABLE", "PARENT_TBL"])
                idx_pc  = self._find_idx(header, ["PARENT_COLUMN", "PARENT_COL"])
                idx_fk  = self._find_idx(header, ["FK_NAME", "FK_CONSTRAINT_NAME"])
                idx_pk  = self._find_idx(header, ["PK_NAME", "PK_CONSTRAINT_NAME"])
                idx_pos = self._find_idx(header, ["POSITION", "POSITION_NO"])

                for r in rows[1:]:
                    if not any(r):
                        continue
                    ct = str(r[idx_ct]).strip() if idx_ct < len(r) and r[idx_ct] else ""
                    cc = str(r[idx_cc]).strip() if idx_cc < len(r) and r[idx_cc] else ""
                    pt = str(r[idx_pt]).strip() if idx_pt < len(r) and r[idx_pt] else ""
                    pc = str(r[idx_pc]).strip() if idx_pc < len(r) and r[idx_pc] else ""
                    if ct and cc and pt and pc:
                        fk = FKConstraint(
                            module_name=str(r[idx_mod]).strip() if idx_mod < len(r) and r[idx_mod] else "Provider",
                            child_table=ct,
                            child_column=cc,
                            parent_table=pt,
                            parent_column=pc,
                            fk_name=str(r[idx_fk]).strip() if idx_fk < len(r) and r[idx_fk] else f"FK_{ct}_{cc}",
                            pk_name=str(r[idx_pk]).strip() if idx_pk < len(r) and r[idx_pk] else "",
                            position_no=int(r[idx_pos]) if idx_pos < len(r) and r[idx_pos] and str(r[idx_pos]).isdigit() else 1,
                        )
                        self.fks.setdefault(ct.lower(), []).append(fk)

        # Load Sheet 2: Primary Keys
        if "Sheet2" in wb.sheetnames:
            ws = wb["Sheet2"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(c).upper() if c else "" for c in rows[0]]
                idx_mod = self._find_idx(header, ["MODULE", "MODULE_NAME"])
                idx_tbl = self._find_idx(header, ["TABLE_NAME", "TABLE"])
                idx_col = self._find_idx(header, ["COLUMN_NAME", "COLUMN"])
                idx_con = self._find_idx(header, ["CONSTRAINT_NAME", "CONSTRAINT"])

                for r in rows[1:]:
                    if not any(r):
                        continue
                    tbl = str(r[idx_tbl]).strip() if idx_tbl < len(r) and r[idx_tbl] else ""
                    col = str(r[idx_col]).strip() if idx_col < len(r) and r[idx_col] else ""
                    con = str(r[idx_con]).strip() if idx_con < len(r) and r[idx_con] else ""
                    if tbl and col:
                        pk = PKConstraint(
                            module_name=str(r[idx_mod]).strip() if idx_mod < len(r) and r[idx_mod] else "Provider",
                            table_name=tbl,
                            column_name=col,
                            constraint_name=con or f"PK_{tbl}",
                        )
                        self.pks.setdefault(tbl.lower(), []).append(pk)

        logger.info("Metadata loaded: %d tables with PKs, %d tables with FKs.", len(self.pks), len(self.fks))

    def _load_from_sql(self, sql_path: Path) -> None:
        logger.info("Loading PK/FK metadata from SQL: %s", sql_path.name)
        text = sql_path.read_text(encoding="utf-8", errors="replace")

        # Parse PK inserts
        pk_re = re.compile(
            r"INSERT\s+INTO\s+migration_primary_keys\s*\([^)]*\)\s*VALUES\s*\(\s*'([^']*)'\s*,\s*'([^']*)'\s*,\s*'([^']*)'\s*,\s*'([^']*)'\s*\)",
            re.IGNORECASE,
        )
        for m in pk_re.finditer(text):
            mod, tbl, col, con = m.group(1), m.group(2), m.group(3), m.group(4)
            pk = PKConstraint(module_name=mod, table_name=tbl, column_name=col, constraint_name=con)
            self.pks.setdefault(tbl.lower(), []).append(pk)

        # Parse FK inserts
        fk_re = re.compile(
            r"INSERT\s+INTO\s+migration_foreign_keys\s*\([^)]*\)\s*VALUES\s*\(\s*'([^']*)'\s*,\s*'([^']*)'\s*,\s*'([^']*)'\s*,\s*'([^']*)'\s*,\s*'([^']*)'\s*,\s*'([^']*)'\s*,\s*'([^']*)'\s*,\s*(\d+|NULL)\s*\)",
            re.IGNORECASE,
        )
        for m in fk_re.finditer(text):
            mod, ct, cc, pt, pc, fk_n, pk_n, pos = m.groups()
            fk = FKConstraint(
                module_name=mod, child_table=ct, child_column=cc,
                parent_table=pt, parent_column=pc, fk_name=fk_n, pk_name=pk_n,
                position_no=int(pos) if pos and pos != "NULL" else 1,
            )
            self.fks.setdefault(ct.lower(), []).append(fk)

        logger.info("Metadata loaded from SQL: %d tables with PKs, %d tables with FKs.", len(self.pks), len(self.fks))

    @staticmethod
    def is_common_table(table_name: str) -> bool:
        """Return True if table belongs to the Common schema."""
        t = table_name.lower().strip()
        return t.startswith("g_") or t.startswith("r_")

    @classmethod
    def get_schema_for_table(cls, table_name: str, default_schema: str = "provider", common_schema: str = "common") -> str:
        """Return the target PostgreSQL schema for a given table name."""
        return common_schema if cls.is_common_table(table_name) else default_schema

    @classmethod
    def get_qualified_table_name(cls, table_name: str, default_schema: str = "provider", common_schema: str = "common") -> str:
        """Return schema-qualified table name (e.g. 'common.g_cmn_enty_tb' or 'provider.p_dtl_tb')."""
        s = cls.get_schema_for_table(table_name, default_schema, common_schema)
        return f"{s}.{table_name.lower().strip()}"

    @staticmethod
    def _find_idx(header: list[str], candidates: list[str]) -> int:
        for cand in candidates:
            if cand in header:
                return header.index(cand)
        return 0
