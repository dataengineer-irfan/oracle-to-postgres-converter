import os
from pathlib import Path
import csv
import psycopg
import re
from datetime import datetime
from openpyxl import load_workbook
import sys

import sys
sys.path.append(os.getcwd())
from backend.config import DB_CONFIG

INPUT_DIR = Path("_Input/Address_Data")
OUTPUT_DIR = Path("input/data")

def clean_oracle_timestamp(val):
    if not isinstance(val, str):
        return val
    # "29-JUL-14 12.35.21.026000000 PM" -> "2014-07-29 12:35:21 PM"
    m = re.match(r"^(\d{1,2}-[A-Za-z]{3}-\d{2})\s+(\d{1,2})\.(\d{2})\.(\d{2})\.\d+\s+(AM|PM)$", val.strip())
    if m:
        dt_str = f"{m.group(1)} {m.group(2)}:{m.group(3)}:{m.group(4)} {m.group(5)}"
        try:
            dt = datetime.strptime(dt_str, "%d-%b-%y %I:%M:%S %p")
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except:
            pass
            
    m2 = re.match(r"^(\d{1,2}-[A-Za-z]{3}-\d{2})$", val.strip())
    if m2:
        try:
            dt = datetime.strptime(m2.group(1), "%d-%b-%y")
            return dt.strftime("%Y-%m-%d")
        except:
            pass
            
    return val

def get_primary_key(conn, schema, table_name):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT kcu.column_name 
            FROM information_schema.table_constraints tco 
            JOIN information_schema.key_column_usage kcu 
              ON kcu.constraint_name = tco.constraint_name 
             AND kcu.constraint_schema = tco.constraint_schema 
            WHERE tco.constraint_type = 'PRIMARY KEY' 
              AND kcu.table_name = %s 
              AND kcu.table_schema = %s;
        """, (table_name, schema))
        row = cur.fetchone()
        return row[0] if row else None

def process_file(conn, file_path):
    table_name = file_path.stem.lower()
    print(f"Processing {table_name} ...")
    
    # Read Excel
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    headers = [cell.value.lower() if cell.value else "" for cell in ws[1]]
    if not headers or not headers[0]:
        print(f"No headers found in {file_path}")
        return
        
    # Write to CSV for the PatternAnalyzer
    csv_path = OUTPUT_DIR / f"{table_name}.csv"
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip completely empty rows
            if not any(row):
                continue
            # Replace Nones with empty strings for CSV
            csv_row = [str(cell) if cell is not None else "" for cell in row]
            writer.writerow(csv_row)
            
            # Keep raw values for DB UPSERT
            rows.append(row)
            
    print(f"Saved {len(rows)} rows to {csv_path}")
    
    # Hardcoded business keys since DB constraints are missing
    keys_map = {
        'g_cmn_enty_tb': ['g_cmn_enty_sk'],
        'g_adr_tb': ['g_adr_sk'],
        'g_adr_usg_tb': ['g_cmn_enty_sk', 'g_adr_sk', 'g_adr_usg_ty_cd']
    }
    
    pk_cols = keys_map.get(table_name)
    if not pk_cols:
        print(f"Unknown PKs for {table_name}")
        return
        
    print(f"Using business keys {pk_cols} for {table_name}")
    
    update_cols = [c for c in headers if c not in pk_cols]
    
    # Execute UPSERT manually
    with conn.cursor() as cur:
        success = 0
        for row in rows:
            db_row = [clean_oracle_timestamp(val) if val != "" else None for val in row]
            row_dict = dict(zip(headers, db_row))
            
            # Check existence
            where_clause = " AND ".join([f"{k}=%s" for k in pk_cols])
            where_vals = [row_dict[k] for k in pk_cols]
            
            cur.execute(f"SELECT 1 FROM common.{table_name} WHERE {where_clause}", where_vals)
            exists = cur.fetchone()
            
            try:
                if exists:
                    # Update
                    if update_cols:
                        set_clause = ", ".join([f"{c}=%s" for c in update_cols])
                        set_vals = [row_dict[c] for c in update_cols]
                        cur.execute(f"UPDATE common.{table_name} SET {set_clause} WHERE {where_clause}", set_vals + where_vals)
                else:
                    # Insert
                    cols = ", ".join(headers)
                    placeholders = ", ".join(["%s"] * len(headers))
                    cur.execute(f"INSERT INTO common.{table_name} ({cols}) VALUES ({placeholders})", db_row)
                success += 1
            except Exception as e:
                print(f"Error on row: {db_row}")
                print(e)
                conn.rollback()
                break
        else:
            conn.commit()
            print(f"Successfully UPSERTed {success} rows into common.{table_name}")

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    try:
        with psycopg.connect(**DB_CONFIG) as conn:
            for file_name in ["G_CMN_ENTY_TB.xlsx", "G_ADR_TB.xlsx", "G_ADR_USG_TB.xlsx"]:
                file_path = INPUT_DIR / file_name
                if file_path.exists():
                    process_file(conn, file_path)
                else:
                    print(f"File not found: {file_path}")
    except Exception as e:
        print(f"Database connection error: {e}")

if __name__ == "__main__":
    main()
